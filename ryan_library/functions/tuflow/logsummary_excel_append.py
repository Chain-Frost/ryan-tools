# ryan_library/functions/tuflow/logsummary_excel_append.py
"""Helpers for appending TUFLOW log-summary rows to an existing Excel table."""

__lazy_modules__ = ["pandas"]


from pathlib import Path
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from decimal import Decimal
import math
from typing import cast

from loguru import logger
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.table import Table, TableColumn
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd

from ryan_library.classes.tuflow_string_classes import TuflowStringParser

ExcelCellValue = str | int | float | bool | datetime | date | time | timedelta | Decimal | None


@dataclass(slots=True, frozen=True)
class ExistingLogSummaryRows:
    """Existing workbook keys used to avoid appending duplicate log-summary rows."""

    headers: tuple[str, ...]
    runcodes: frozenset[str]
    log_paths: frozenset[str]
    log_filenames: frozenset[str]


def load_existing_log_summary_rows(
    *,
    workbook_path: Path,
    sheet_name: str,
    table_name: str | None,
) -> ExistingLogSummaryRows:
    """Read existing run-code and log-path keys from a log-summary workbook table."""
    if not workbook_path.exists():
        raise FileNotFoundError(f"Master workbook does not exist: {workbook_path}")

    workbook: Workbook = load_workbook(filename=workbook_path, read_only=False, data_only=False)
    try:
        worksheet: Worksheet = _resolve_worksheet(workbook=workbook, sheet_name=sheet_name)
        table: Table | None = _resolve_table(worksheet=worksheet, table_name=table_name)
        min_col, header_row, max_col, max_row = _table_or_sheet_bounds(worksheet=worksheet, table=table)
        headers: tuple[str, ...] = _read_header_row(
            worksheet=worksheet,
            header_row=header_row,
            min_col=min_col,
            max_col=max_col,
        )
        runcode_index: int | None = _column_offset(headers=headers, column_name="Runcode")
        log_path_index: int | None = _column_offset(headers=headers, column_name="orig_log_path")

        runcodes: set[str] = set()
        log_paths: set[str] = set()
        log_filenames: set[str] = set()
        for row in worksheet.iter_rows(
            min_row=header_row + 1,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        ):
            if runcode_index is not None and runcode_index < len(row):
                if runcode_key := _normalise_text_key(row[runcode_index]):
                    runcodes.add(runcode_key)

            if log_path_index is not None and log_path_index < len(row):
                raw_log_path: object = row[log_path_index]
                if path_key := _normalise_path_key(raw_log_path):
                    log_paths.add(path_key)
                if filename_key := _normalise_path_filename_key(raw_log_path):
                    log_filenames.add(filename_key)

        logger.info(f"Loaded {len(runcodes)} existing run code(s) from {workbook_path}")
        return ExistingLogSummaryRows(
            headers=headers,
            runcodes=frozenset(runcodes),
            log_paths=frozenset(log_paths),
            log_filenames=frozenset(log_filenames),
        )
    finally:
        workbook.close()


def filter_new_log_files(
    *,
    files: Iterable[Path],
    existing_rows: ExistingLogSummaryRows,
) -> list[Path]:
    """Return only log files whose run code/path is not already in the workbook."""
    new_files: list[Path] = []
    for logfile in files:
        runcode_key: str | None = _runcode_key_from_logfile(logfile=logfile)
        path_key: str | None = _normalise_path_key(logfile)
        filename_key: str | None = _normalise_path_filename_key(logfile)

        if runcode_key and runcode_key in existing_rows.runcodes:
            logger.debug("Skipping existing run code {}", logfile)
            continue
        if path_key and path_key in existing_rows.log_paths:
            logger.debug("Skipping existing log path {}", logfile)
            continue
        if filename_key and filename_key in existing_rows.log_filenames:
            logger.debug("Skipping existing log filename {}", logfile)
            continue

        new_files.append(logfile)
    return new_files


def append_dataframe_to_workbook_table(
    *,
    data_frame: pd.DataFrame,
    workbook_path: Path,
    sheet_name: str,
    table_name: str | None,
) -> None:
    """Append dataframe rows to a worksheet table, extending the table range as needed."""
    workbook: Workbook = load_workbook(filename=workbook_path, read_only=False, data_only=False)
    try:
        worksheet: Worksheet = _resolve_worksheet(workbook=workbook, sheet_name=sheet_name)
        table: Table | None = _resolve_table(worksheet=worksheet, table_name=table_name)
        min_col, header_row, max_col, max_row = _table_or_sheet_bounds(worksheet=worksheet, table=table)
        headers: list[str] = list(
            _read_header_row(
                worksheet=worksheet,
                header_row=header_row,
                min_col=min_col,
                max_col=max_col,
            )
        )

        _add_missing_headers(
            worksheet=worksheet,
            data_frame=data_frame,
            headers=headers,
            header_row=header_row,
            min_col=min_col,
        )
        max_col = min_col + len(headers) - 1

        append_start_row: int = max(max_row + 1, header_row + 1)
        for row_offset, (_, series) in enumerate(data_frame.iterrows(), start=append_start_row):
            for column_offset, column_name in enumerate(headers, start=min_col):
                worksheet.cell(
                    row=row_offset,
                    column=column_offset,
                    value=_excel_safe_value(series.get(column_name)),
                )

        append_end_row: int = append_start_row + len(data_frame.index) - 1
        if table is not None:
            table.ref = _excel_range_ref(
                min_col=min_col,
                min_row=header_row,
                max_col=max_col,
                max_row=append_end_row,
            )
            _sync_table_columns(table=table, headers=headers)
            if table.autoFilter is not None:
                table.autoFilter.ref = table.ref

        logger.info(f"Saving appended workbook {workbook_path}")
        workbook.save(filename=workbook_path)
    except PermissionError:
        logger.error(f"Unable to save {workbook_path}. Close it in Excel and rerun the append script.")
        raise
    finally:
        workbook.close()


def _resolve_worksheet(*, workbook: Workbook, sheet_name: str) -> Worksheet:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}")
    return workbook[sheet_name]


def _resolve_table(*, worksheet: Worksheet, table_name: str | None) -> Table | None:
    table_names: list[str] = list(worksheet.tables.keys())
    if not table_names:
        logger.warning(f"Worksheet '{worksheet.title}' has no Excel table; appending to used range instead.")
        return None

    resolved_table_name: str = table_name or table_names[0]
    if resolved_table_name not in worksheet.tables:
        raise ValueError(
            f"Table '{resolved_table_name}' not found on sheet '{worksheet.title}'. " f"Available tables: {table_names}"
        )
    table = worksheet.tables[resolved_table_name]
    if not isinstance(table, Table):
        raise TypeError(f"Expected openpyxl Table for '{resolved_table_name}', got {type(table).__name__}")
    return table


def _table_or_sheet_bounds(*, worksheet: Worksheet, table: Table | None) -> tuple[int, int, int, int]:
    if table is not None:
        bounds = range_boundaries(table.ref)
        if any(bound is None for bound in bounds):
            raise ValueError(f"Invalid table range: {table.ref}")
        return cast(tuple[int, int, int, int], bounds)
    return 1, 1, worksheet.max_column, worksheet.max_row


def _read_header_row(
    *,
    worksheet: Worksheet,
    header_row: int,
    min_col: int,
    max_col: int,
) -> tuple[str, ...]:
    headers: list[str] = []
    for cell in worksheet.iter_rows(
        min_row=header_row,
        max_row=header_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True,
    ):
        headers = ["" if value is None else str(value) for value in cell]
        break
    if not headers or all(not header for header in headers):
        raise ValueError(f"No headers found on worksheet '{worksheet.title}' row {header_row}.")
    return tuple(headers)


def _column_offset(*, headers: tuple[str, ...], column_name: str) -> int | None:
    lowered_column_name: str = column_name.lower()
    for index, header in enumerate(headers):
        if header.lower() == lowered_column_name:
            return index
    return None


def _add_missing_headers(
    *,
    worksheet: Worksheet,
    data_frame: pd.DataFrame,
    headers: list[str],
    header_row: int,
    min_col: int,
) -> None:
    existing_headers: set[str] = set(headers)
    for column_name in data_frame.columns:
        column_name_text: str = str(column_name)
        if column_name_text in existing_headers:
            continue
        headers.append(column_name_text)
        existing_headers.add(column_name_text)
        worksheet.cell(row=header_row, column=min_col + len(headers) - 1, value=column_name_text)
        logger.info(f"Added new master log column: {column_name_text}")


def _sync_table_columns(*, table: Table, headers: list[str]) -> None:
    while len(table.tableColumns) < len(headers):
        column_position: int = len(table.tableColumns) + 1
        table.tableColumns.append(TableColumn(id=column_position, name=headers[column_position - 1]))

    for column_position, header in enumerate(headers, start=1):
        table_column: TableColumn = table.tableColumns[column_position - 1]
        table_column.id = column_position
        table_column.name = header


def _excel_range_ref(*, min_col: int, min_row: int, max_col: int, max_row: int) -> str:
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def _excel_safe_value(value: object) -> ExcelCellValue:
    if value is None:
        return None

    if value is pd.NA or value is pd.NaT:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, Decimal) and value.is_nan():
        return None

    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, (str, int, float, bool, datetime, date, time, timedelta, Decimal)):
        return value

    item_method = getattr(value, "item", None)
    if callable(item_method):
        try:
            return _excel_safe_value(item_method())
        except TypeError, ValueError:
            pass

    return str(value)


def _runcode_key_from_logfile(*, logfile: Path) -> str | None:
    try:
        return _normalise_text_key(TuflowStringParser(file_path=logfile).clean_run_code)
    except Exception as exc:
        logger.debug("Unable to parse run code from {}: {}", logfile, exc)
        return _normalise_text_key(logfile.stem)


def _normalise_text_key(value: object) -> str | None:
    if value is None:
        return None
    text: str = str(value).strip()
    if not text:
        return None
    lowered: str = text.lower()
    if lowered in {"nan", "none", "<na>"}:
        return None
    return lowered


def _normalise_path_key(value: object) -> str | None:
    text_key: str | None = _normalise_text_key(value)
    if text_key is None:
        return None
    return text_key.replace("/", "\\")


def _normalise_path_filename_key(value: object) -> str | None:
    path_key: str | None = _normalise_path_key(value)
    if path_key is None:
        return None
    return path_key.rsplit("\\", 1)[-1].lower()
