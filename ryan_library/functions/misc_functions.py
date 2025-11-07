# ryan_library/functions/misc_functions.py

from datetime import datetime
import multiprocessing
import pandas as pd
import logging
from typing import TypedDict
from pathlib import Path
from importlib import metadata
import re
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.exceptions import InvalidFileException
from ryan_library.functions.logging_helpers import setup_logging as new_setup_logging


def get_tools_version(package: str = "ryan_functions") -> str:
    """Return the installed version of ``package`` if available."""
    try:
        return metadata.version(distribution_name=package)
    except metadata.PackageNotFoundError:
        return "unknown"


# Deprecated setup_logging function
def setup_logging(
    log_level: int = logging.INFO,
    log_file: str | None = None,
    max_bytes: int = 10**6,  # 1 MB
    backup_count: int = 5,
    use_rotating_file: bool = False,
    enable_color: bool = True,
) -> None:
    print("This is using a deprecated logging procedure")
    new_setup_logging(
        log_level=log_level,
        log_file=log_file,
        max_bytes=max_bytes,
        backup_count=backup_count,
        use_rotating_file=use_rotating_file,
        enable_color=enable_color,
    )


def calculate_pool_size(num_files: int) -> int:
    """Calculate the optimal pool size based on the number of files and CPU cores.
    Args:
        num_files (int): Number of files to process.
    Returns:
        int: Number of threads to use."""
    splits: int = max(num_files // 3, 1)
    available_cores: int = min(multiprocessing.cpu_count(), 20)
    calc_threads: int = min(available_cores - 1, splits) if available_cores > 1 else 1
    logging.info(f"Processing threads: {calc_threads}")
    return calc_threads


def split_strings(input_str: str | list[str]) -> list[str]:
    """Split input string(s) by whitespace into a flat list of strings.
    Args:
        input_str (str | list[str]): A string or list of strings to split.
    Returns:
        list[str]: A flat list of split strings."""
    if isinstance(input_str, str):
        input_list: list[str] = [input_str]
    else:  # input is already a list
        input_list = input_str

    # Split each string by whitespace and flatten the list
    split_list: list[str] = []
    for item in input_list:
        split_list.extend(item.split())

    return split_list


def split_strings_in_dict(params_dict: dict[str, list[str]]) -> dict[str, list[str]]:
    """Apply split_strings to each list of strings in the dictionary.
    Args:
        params_dict (dict[str, list[str]]): Dictionary with string lists to split.
    Returns:
        dict[str, list[str]]: Dictionary with split string lists."""
    for key, value in params_dict.items():
        # Use split_strings to handle both string and list of strings cases
        params_dict[key] = split_strings(input_str=value)
    return params_dict


class ExportContent(TypedDict):
    dataframes: list[pd.DataFrame]
    sheets: list[str]


class ExcelExporter:
    """A utility class for exporting pandas DataFrames to Excel files.

    Methods:
        export_dataframes: Export multiple DataFrames to Excel with optional column widths.
        save_to_excel: Export a single DataFrame to Excel with optional column widths.
        calculate_column_widths: Calculate optimal column widths based on data.
        set_column_widths: Apply specific column widths to a worksheet.
        auto_adjust_column_widths: Automatically adjust column widths based on data."""

    MAX_EXCEL_ROWS: int = 1_048_576
    MAX_EXCEL_COLUMNS: int = 16_384

    def export_dataframes(
        self,
        export_dict: dict[str, ExportContent],
        output_directory: Path | None = None,
        column_widths: dict[str, dict[str, float]] | None = None,
        auto_adjust_width: bool = True,
    ) -> None:
        """Export multiple DataFrames to Excel files with optional column widths.
        Args:
            export_dict (dict[str, ExportContent]):
                A dictionary where each key is a base file name and each value contains:
                    - "dataframes": list of DataFrames
                    - "sheets": list of corresponding sheet names
            output_directory (Path | None, optional):
                The directory where the Excel files will be saved.
                Defaults to the current working directory.
            column_widths (dict[str, dict[str, float]] | None, optional):
                A dictionary where keys are sheet names and values are dictionaries
                mapping column names to their desired widths.
                Example:
                    {
                        "Sheet1": {"Name": 20, "Age": 10},
                        "Sheet2": {"Email": 30}
                    }
            auto_adjust_width (bool, optional):
                If set to True, automatically adjusts the column widths based on the
                maximum length of the data in each column. Defaults to True.
        Raises:
            ValueError: If the number of DataFrames doesn't match the number of sheets.
            InvalidFileException: If there's an issue with writing the Excel file.
        Example:
            export_dict = {
                "Report": {
                    "dataframes": [df1, df2],
                    "sheets": ["Summary", "Details"]
                },
                "AnotherReport": {
                    "dataframes": [df3],
                    "sheets": ["Data"]
                }
            }
            ExcelExporter().export_dataframes(export_dict, output_directory=Path("exports"))
        """
        datetime_string: str = datetime.now().strftime(format="%Y%m%d-%H%M")

        for file_name, content in export_dict.items():
            dataframes: list[pd.DataFrame] = content.get("dataframes", [])
            sheets: list[str] = content.get("sheets", [])

            if len(dataframes) != len(sheets):
                raise ValueError(
                    f"For file '{file_name}', the number of dataframes ({len(dataframes)}) and sheets ({len(sheets)}) must match."
                )

            if self._exceeds_excel_limits(dataframes=dataframes):
                logging.warning(
                    "Data for '%s' exceeds Excel size limits. Exporting to Parquet and CSV instead.",
                    file_name,
                )
                self._export_as_parquet_and_csv(
                    file_name=file_name,
                    dataframes=dataframes,
                    sheets=sheets,
                    datetime_string=datetime_string,
                    output_directory=output_directory,
                )
                continue

            # Determine the export path
            export_filename: str = f"{datetime_string}_{file_name}.xlsx"
            export_path: Path = (
                (output_directory / export_filename) if output_directory else Path(export_filename)  # Defaults to CWD
            )

            # Ensure the output directory exists
            if output_directory:
                export_path.parent.mkdir(parents=True, exist_ok=True)

            logging.info(f"Exporting to {export_path}")

            try:
                with pd.ExcelWriter(path=export_path, engine="openpyxl") as writer:
                    writer.book.properties.creator = f"ryan-tools {get_tools_version()}"
                    for df, sheet in zip(dataframes, sheets):
                        # Check for unique column names
                        if not df.columns.is_unique:
                            logging.error(
                                f"DataFrame for sheet '{sheet}' has duplicate column names. Please ensure all column names are unique."
                            )
                            raise ValueError(f"Duplicate column names found in sheet '{sheet}'.")

                        df.to_excel(
                            excel_writer=writer,
                            sheet_name=sheet,
                            merge_cells=False,
                            index=False,
                        )

                        # Access the worksheet
                        workbook: Workbook = writer.book
                        worksheet: Worksheet = writer.sheets[sheet]

                        # Automatically adjust column widths if enabled
                        if auto_adjust_width:
                            dynamic_widths: dict[str, float] = self.calculate_column_widths(df=df)
                            self.auto_adjust_column_widths(worksheet=worksheet, dynamic_widths=dynamic_widths)

                        # Apply specific column widths if provided
                        if column_widths and sheet in column_widths:
                            self.set_column_widths(
                                worksheet=worksheet,
                                df=df,
                                sheet_name=sheet,
                                column_widths=column_widths[sheet],
                            )

                logging.info(f"Finished exporting '{file_name}' to '{export_path}'")
            except InvalidFileException as e:
                logging.error(f"Failed to write to '{export_path}': {e}")
                raise

    def _exceeds_excel_limits(self, dataframes: list[pd.DataFrame]) -> bool:
        """Return True if any dataframe exceeds Excel's size limits."""

        for df in dataframes:
            num_data_rows: int = len(df.index)
            num_columns: int = len(df.columns)
            header_rows: int = df.columns.nlevels if num_columns > 0 else 0
            total_rows: int = num_data_rows + header_rows

            if total_rows > self.MAX_EXCEL_ROWS or num_columns > self.MAX_EXCEL_COLUMNS:
                logging.debug(
                    "Dataframe size rows=%s (including %s header rows) columns=%s exceeds Excel limits (rows=%s, columns=%s).",
                    total_rows,
                    header_rows,
                    num_columns,
                    self.MAX_EXCEL_ROWS,
                    self.MAX_EXCEL_COLUMNS,
                )
                return True
        return False

    def _export_as_parquet_and_csv(
        self,
        file_name: str,
        dataframes: list[pd.DataFrame],
        sheets: list[str],
        datetime_string: str,
        output_directory: Path | None,
    ) -> None:
        """Export dataframes to Parquet and CSV files when Excel limits are exceeded."""

        export_targets: list[tuple[pd.DataFrame, str, Path, Path]] = []

        for df, sheet in zip(dataframes, sheets):
            sanitized_sheet: str = self._sanitize_name(sheet)
            base_filename: str = f"{datetime_string}_{file_name}_{sanitized_sheet}"

            parquet_path: Path = self._build_output_path(
                base_filename=f"{base_filename}.parquet", output_directory=output_directory
            )
            csv_path: Path = self._build_output_path(
                base_filename=f"{base_filename}.csv", output_directory=output_directory
            )

            parquet_path.parent.mkdir(parents=True, exist_ok=True)
            export_targets.append((df, sheet, parquet_path, csv_path))

        for df, sheet, parquet_path, _ in export_targets:
            try:
                df.to_parquet(path=parquet_path, index=False)
                logging.info("Exported Parquet to %s", parquet_path)
            except (ImportError, ValueError) as exc:
                message: str = (
                    "Unable to export Parquet for "
                    f"'{file_name}' sheet '{sheet}': {exc}. Install pyarrow or fastparquet."
                )
                logging.error(message)
                print(message)
            except Exception as exc:  # pragma: no cover - unforeseen errors should be logged
                logging.exception(
                    "Unexpected error during Parquet export for '%s' sheet '%s': %s", file_name, sheet, exc
                )

        for df, sheet, _, csv_path in export_targets:
            df.to_csv(path_or_buf=csv_path, index=False)
            logging.info("Exported CSV to %s", csv_path)

    def _build_output_path(self, base_filename: str, output_directory: Path | None) -> Path:
        """Create the full output path for a file name."""

        if output_directory is not None:
            return output_directory / base_filename
        return Path(base_filename)

    def _sanitize_name(self, value: str) -> str:
        """Return a filesystem-friendly version of the provided value."""

        sanitized: str = re.sub(pattern=r"[^A-Za-z0-9_-]+", repl="_", string=value).strip("_")
        return sanitized or "Sheet"

    def save_to_excel(
        self,
        data_frame: pd.DataFrame,
        file_name_prefix: str = "Export",
        sheet_name: str = "Export",
        output_directory: Path | None = None,
        column_widths: dict[str, float] | None = None,
        auto_adjust_width: bool = True,
    ) -> None:
        """Export a single DataFrame to an Excel file with a single sheet and optional column widths.

        Args:
            data_frame (pd.DataFrame): The DataFrame to export.
            file_name_prefix (str): Prefix for the resulting Excel filename.
            sheet_name (str): Name of the sheet in the Excel file.
            output_directory (Path | None, optional):
                The directory where the Excel file will be saved.
                Defaults to the current working directory.
            column_widths (dict[str, float] | None, optional):
                A dictionary mapping column names to their desired widths.
                Example:
                    {"Name": 20, "Age": 10}
            auto_adjust_width (bool, optional):
                If set to True, automatically adjusts the column widths based on the
                maximum length of the data in each column. Defaults to True."""
        export_dict: dict[str, ExportContent] = {file_name_prefix: {"dataframes": [data_frame], "sheets": [sheet_name]}}

        # Prepare column_widths in the required format
        prepared_column_widths: dict[str, dict[str, float]] | None = (
            {sheet_name: column_widths} if column_widths else None
        )

        self.export_dataframes(
            export_dict=export_dict,
            output_directory=output_directory,
            column_widths=prepared_column_widths,
            auto_adjust_width=auto_adjust_width,
        )

    def calculate_column_widths(self, df: pd.DataFrame) -> dict[str, float]:
        """Calculate optimal column widths based on the maximum length of data in each column.

        Args:
            df (pd.DataFrame): The DataFrame for which to calculate column widths.

        Returns:
            dict[str, float]: A dictionary mapping column letters to their calculated widths.
        """
        column_widths: dict[str, float] = {
            get_column_letter(idx + 1): max(
                # Calculate max length of the column data
                df[col].astype(str).map(len).max(),
                # Consider the column name length as well
                len(str(col)),
            )
            + 2  # Adding extra space
            for idx, col in enumerate(df.columns)
        }
        logging.debug(f"Calculated dynamic column widths: {column_widths}")
        return column_widths

    def set_column_widths(
        self,
        worksheet: Worksheet,
        df: pd.DataFrame,
        sheet_name: str,
        column_widths: dict[str, float],
    ) -> None:
        """Set specific column widths for a given worksheet based on provided configurations.
        Args:
            worksheet (Worksheet): The OpenPyXL worksheet object.
            df (pd.DataFrame): The pandas DataFrame containing the data.
            sheet_name (str): The name of the current sheet.
            column_widths (dict[str, float]):
                A dictionary mapping column names to their desired widths.
        Raises:
            TypeError: If column indices are not integers."""
        for col_name, width in column_widths.items():
            if col_name not in df.columns:
                logging.warning(f"Column '{col_name}' not found in sheet '{sheet_name}'. Skipping width setting.")
                continue

            try:

                col_idx = df.columns.get_loc(col_name)
                assert isinstance(col_idx, int), (
                    f"Expected integer column index for '{col_name}' in sheet '{sheet_name}', "
                    f"but got {type(col_idx).__name__}"
                )
                col_letter: str = get_column_letter(col_idx + 1)
                worksheet.column_dimensions[col_letter].width = width
                logging.debug(f"Set width for column '{col_name}' ({col_letter}) in sheet '{sheet_name}' to {width}.")
            except TypeError as e:
                logging.exception(f"TypeError when setting width for column '{col_name}' in sheet '{sheet_name}': {e}")
            except AssertionError as e:
                logging.exception(str(e))
            except Exception as e:
                logging.exception(
                    f"Unexpected error when setting width for column '{col_name}' in sheet '{sheet_name}': {e}"
                )

    def auto_adjust_column_widths(self, worksheet: Worksheet, dynamic_widths: dict[str, float]) -> None:
        """Automatically adjust column widths based on calculated dynamic widths.
        Args:
            worksheet (Worksheet): The OpenPyXL worksheet object.
            dynamic_widths (dict[str, float]): Calculated dynamic column widths."""
        for col_letter, width in dynamic_widths.items():
            current_width: float | None = worksheet.column_dimensions[col_letter].width
            if current_width is None or width > current_width:
                worksheet.column_dimensions[col_letter].width = width
                logging.debug(f"Auto-adjusted width for column '{col_letter}' to {width}.")


# Backwards compatibility functions:
def export_dataframes(export_dict: dict[str, ExportContent], output_directory: Path | None = None) -> None:
    """Backwards-compatible function that delegates to ExcelExporter.
    Args:
        export_dict (dict[str, ExportContent]):
            Dictionary containing export information.
        output_directory (Path, optional):
            Directory to save the exported Excel files.
            Defaults to the current working directory."""
    ExcelExporter().export_dataframes(export_dict=export_dict, output_directory=output_directory)


def save_to_excel(
    data_frame: pd.DataFrame,
    file_name_prefix: str = "Export",
    sheet_name: str = "Export",
    output_directory: Path | None = None,
) -> None:
    """Backwards-compatible function that delegates to ExcelExporter.
    Args:
        data_frame (pd.DataFrame): The DataFrame to export.
        file_name_prefix (str): Prefix for the resulting Excel filename.
        sheet_name (str): Name of the sheet in the Excel file.
        output_directory (Path, optional):
            Directory to save the exported Excel file.
            Defaults to the current working directory."""
    ExcelExporter().save_to_excel(
        data_frame=data_frame,
        file_name_prefix=file_name_prefix,
        sheet_name=sheet_name,
        output_directory=output_directory,
    )
