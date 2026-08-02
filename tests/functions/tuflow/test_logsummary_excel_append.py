"""Tests for ryan_library.functions.tuflow.logsummary_excel_append."""

import pytest
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from ryan_library.functions.tuflow import logsummary_excel_append as lsea


def create_dummy_workbook(path: Path, with_table: bool = True) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    # Add headers
    ws.append(["Runcode", "orig_log_path", "Other"])
    # Add data
    ws.append(["run1", "path/to/run1.tlf", "data1"])
    ws.append(["run2", "path/to/run2.tlf", "data2"])

    if with_table:
        tab = Table(displayName="Table1", ref="A1:C3")
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)

    wb.save(path)


class TestExistingLogSummaryRows:
    def test_load_existing_log_summary_rows_with_table(self, tmp_path: Path) -> None:
        wb_path = tmp_path / "test.xlsx"
        create_dummy_workbook(wb_path, with_table=True)

        result = lsea.load_existing_log_summary_rows(workbook_path=wb_path, sheet_name="Summary", table_name="Table1")
        assert result.headers == ("Runcode", "orig_log_path", "Other")
        assert "run1" in result.runcodes
        assert "run2" in result.runcodes
        assert str(Path("path/to/run1.tlf")) in result.log_paths

    def test_load_existing_log_summary_rows_without_table(self, tmp_path: Path) -> None:
        wb_path = tmp_path / "test_no_table.xlsx"
        create_dummy_workbook(wb_path, with_table=False)

        result = lsea.load_existing_log_summary_rows(workbook_path=wb_path, sheet_name="Summary", table_name=None)
        assert result.headers == ("Runcode", "orig_log_path", "Other")
        assert "run1" in result.runcodes
        assert str(Path("path/to/run2.tlf")) in result.log_paths

    def test_filter_new_log_files(self) -> None:
        existing = lsea.ExistingLogSummaryRows(
            headers=("Runcode", "orig_log_path"),
            runcodes=frozenset(["run1"]),
            log_paths=frozenset([str(Path("C:/logs/run2.tlf").resolve()).lower()]),
            log_filenames=frozenset(["run3.tlf"]),
        )
        files = [
            Path("C:/logs/run1.tlf"),  # runcode matches (run1)
            Path("C:/logs/run2.tlf"),  # path matches (c:/logs/run2.tlf)
            Path("C:/logs/other/run3.tlf"),  # filename matches (run3.tlf)
            Path("C:/logs/run4.tlf"),  # Should be kept
        ]

        filtered = lsea.filter_new_log_files(files=files, existing_rows=existing)
        assert len(filtered) == 1
        assert filtered[0].name == "run4.tlf"


class TestAppendDataframe:
    def test_append_dataframe_to_workbook_table(self, tmp_path: Path) -> None:
        wb_path = tmp_path / "append_test.xlsx"
        create_dummy_workbook(wb_path, with_table=True)

        df = pd.DataFrame({"Runcode": ["run3"], "orig_log_path": ["path/to/run3.tlf"], "NewCol": ["new_data"]})

        lsea.append_dataframe_to_workbook_table(
            data_frame=df, workbook_path=wb_path, sheet_name="Summary", table_name=None
        )

        # Verify
        from openpyxl import load_workbook

        wb = load_workbook(wb_path)
        ws = wb["Summary"]
        headers = [cell.value for cell in ws[1]]
        assert "NewCol" in headers
        assert ws.max_row == 4
        assert ws.cell(row=4, column=1).value == "run3"
        assert ws.cell(row=4, column=headers.index("NewCol") + 1).value == "new_data"
